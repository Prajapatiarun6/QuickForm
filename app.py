import json
import os
import re
import secrets
import sqlite3
from flask import Flask, render_template, request, redirect, jsonify
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

app = Flask(__name__)
DB_FILE = "app_data.db"


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS forms (
            form_id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            fields TEXT NOT NULL,
            filename TEXT NOT NULL,
            id_field TEXT,
            status TEXT DEFAULT 'active'
        )
    """
    )
    conn.commit()
    conn.close()


init_db()


def auto_fit_columns(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        wb.save(file_path)
    except Exception:
        pass


def get_primary_identifier(fields, excel_file=None):
    # Strictly prioritize unique keys in order
    for key in ["email", "phone", "mobile", "roll", "enroll", "name"]:
        for f in fields:
            if key in f.lower():
                return f

    if excel_file and os.path.exists(excel_file):
        try:
            df = pd.read_excel(excel_file, dtype=str)
            existing_cols = [str(c).strip() for c in df.columns]
            if existing_cols:
                return existing_cols[0]
        except Exception:
            pass

    return fields[0] if fields else None


@app.route("/")
def admin_panel():
    conn = get_db()
    rows = conn.execute("SELECT * FROM forms").fetchall()
    conn.close()

    existing_forms = {}
    for r in rows:
        existing_forms[r["form_id"]] = {
            "title": r["title"],
            "fields": json.loads(r["fields"]),
            "filename": r["filename"],
            "id_field": r["id_field"],
            "status": r["status"],
        }
    return render_template("admin.html", view="home", existing_forms=existing_forms)


@app.route("/create-page")
def create_page():
    return render_template("admin.html", view="create", is_edit=False)


@app.route("/edit/<form_id>")
def edit_form(form_id):
    conn = get_db()
    r = conn.execute("SELECT * FROM forms WHERE form_id = ?", (form_id,)).fetchone()
    conn.close()

    if r:
        return render_template(
            "admin.html",
            view="edit",
            is_edit=True,
            form_id=form_id,
            edit_title=r["title"],
            edit_fields=json.loads(r["fields"]),
        )
    return redirect("/")


@app.route("/create-form", methods=["POST"])
def create_form():
    title = request.form.get("form_title", "FormData").strip()
    clean_title = re.sub(r"\s+", "", title)
    clean_title = re.sub(r"\W+", "", clean_title) or "FormData"

    existing_form_id = request.form.get("existing_form_id", "").strip()

    if existing_form_id:
        form_id = existing_form_id
    else:
        unique_suffix = secrets.token_hex(2)
        form_id = f"{clean_title}_{unique_suffix}"

    raw_fields = request.form.getlist("custom_fields[]")
    clean_fields = [f.strip() for f in raw_fields if f.strip() != ""]

    excel_filename = f"{form_id}.xlsx"
    id_field = get_primary_identifier(clean_fields, excel_filename)

    # Force Primary unique field to index 0 (Top position)
    if id_field and id_field in clean_fields:
        clean_fields.remove(id_field)
        clean_fields.insert(0, id_field)

    conn = get_db()
    conn.execute(
        """
        INSERT OR REPLACE INTO forms (form_id, title, fields, filename, id_field, status)
        VALUES (?, ?, ?, ?, ?, COALESCE((SELECT status FROM forms WHERE form_id=?), 'active'))
    """,
        (form_id, clean_title, json.dumps(clean_fields), excel_filename, id_field, form_id),
    )
    conn.commit()
    conn.close()

    if os.path.exists(excel_filename):
        try:
            df = pd.read_excel(excel_filename, dtype=str).fillna("")
            for col in clean_fields:
                if col not in df.columns:
                    df[col] = ""
            df = df[clean_fields]
        except Exception:
            df = pd.DataFrame(columns=clean_fields)
    else:
        df = pd.DataFrame(columns=clean_fields)

    try:
        df.to_excel(excel_filename, index=False)
        auto_fit_columns(excel_filename)
    except PermissionError:
        return "<h2 style='text-align: center; color: red;'>⚠️ Please close Excel on your laptop before saving!</h2>", 500

    form_url = f"{request.host_url}form/{form_id}"

    return f"""
    <div style="font-family: Arial, sans-serif; text-align: center; padding: 40px;">
        <h2>✅ Form Saved Successfully!</h2>
        <p>Form Title: <b>{clean_title}</b></p>
        <p>Shareable Link:</p>
        <p><a href="{form_url}" target="_blank" style="font-size: 18px; color: #1f4e78; font-weight: bold;">{form_url}</a></p>
        <br><br>
        <a href="/" style="background: #1f4e78; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">🏠 Back to Dashboard</a>
    </div>
    """


@app.route("/toggle-status/<form_id>", methods=["POST"])
def toggle_status(form_id):
    conn = get_db()
    r = conn.execute("SELECT status FROM forms WHERE form_id = ?", (form_id,)).fetchone()
    if r:
        new_status = "closed" if r["status"] == "active" else "active"
        conn.execute("UPDATE forms SET status = ? WHERE form_id = ?", (new_status, form_id))
        conn.commit()
    conn.close()
    return redirect("/")


@app.route("/delete-form/<form_id>", methods=["POST"])
def delete_form(form_id):
    conn = get_db()
    conn.execute("DELETE FROM forms WHERE form_id = ?", (form_id,))
    conn.commit()
    conn.close()
    return redirect("/")


@app.route("/get-data/<form_id>")
def get_student_data(form_id):
    query_val = request.args.get("value", "").strip().lower()
    conn = get_db()
    r = conn.execute("SELECT * FROM forms WHERE form_id = ?", (form_id,)).fetchone()
    conn.close()

    if r and r["status"] == "active" and query_val:
        excel_file = r["filename"]
        id_field = r["id_field"]
        if os.path.exists(excel_file) and id_field:
            try:
                df = pd.read_excel(excel_file, dtype=str).fillna("")
                match_col = next(
                    (col for col in df.columns if id_field.lower() in str(col).lower()),
                    None,
                )
                if match_col:
                    match = df[df[match_col].astype(str).str.strip().str.lower() == query_val]
                    if not match.empty:
                        row_data = match.iloc[0].to_dict()
                        return jsonify({"found": True, "data": row_data})
            except Exception:
                pass
    return jsonify({"found": False})


@app.route("/form/<form_id>", methods=["GET", "POST"])
def student_form(form_id):
    conn = get_db()
    r = conn.execute("SELECT * FROM forms WHERE form_id = ?", (form_id,)).fetchone()
    conn.close()

    if not r:
        return "<h2 style='text-align:center;'>❌ Form Not Found!</h2>", 404

    if r["status"] != "active":
        return """
        <div style="font-family: Arial, sans-serif; text-align: center; padding: 50px;">
            <h2 style='color: #d9534f;'>🔒 Submissions Closed</h2>
            <p style='color: #666;'>This form is no longer accepting responses.</p>
        </div>
        """, 403

    fields = json.loads(r["fields"])
    title = r["title"]
    excel_file = r["filename"]
    id_field = get_primary_identifier(fields, excel_file)

    # Force Primary Identifier at top position
    if id_field and id_field in fields:
        fields.remove(id_field)
        fields.insert(0, id_field)

    if request.method == "POST":
        submitted_data = {}
        for idx, field_name in enumerate(fields):
            val = request.form.get(f"field_{idx}", "").strip()
            if "name" in field_name.lower():
                val = val.title()
            submitted_data[field_name] = str(val)

        for key, value in submitted_data.items():
            if "email" in key.lower() and value:
                if not re.match(r"^[^@]+@[^@]+\.[^@]+$", value):
                    return "<h2 style='color: red; text-align: center;'>❌ Invalid Email Address! Please go back and correct it.</h2>"

        id_val = submitted_data.get(id_field, "").strip().lower() if id_field else None

        if os.path.exists(excel_file):
            try:
                df = pd.read_excel(excel_file, dtype=str).fillna("")
            except Exception:
                df = pd.DataFrame(columns=fields)

            df = df.astype(str)

            match_col = next(
                (col for col in df.columns if id_field and id_field.lower() in str(col).lower()),
                None,
            )
            if match_col and id_val and id_val in df[match_col].astype(str).str.strip().str.lower().values:
                row_idx = df[df[match_col].astype(str).str.strip().str.lower() == id_val].index[0]
                for key, val in submitted_data.items():
                    if val:
                        df.at[row_idx, key] = str(val)
            else:
                new_row = pd.DataFrame([submitted_data]).astype(str)
                df = pd.concat([df, new_row], ignore_index=True)
        else:
            df = pd.DataFrame([submitted_data]).astype(str)

        try:
            df.to_excel(excel_file, index=False)
            auto_fit_columns(excel_file)
        except PermissionError:
            return "<h2 style='text-align: center; color: #d9534f;'>⚠️ System Busy: Excel file is open on Admin laptop. Please close it and retry.</h2>", 500

        return "<h2 style='text-align: center; color: green;'>✅ Response Submitted Successfully!</h2>"

    return render_template(
        "index.html",
        form_title=title,
        fields=fields,
        form_id=form_id,
        id_field=id_field,
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)